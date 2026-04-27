import csv
import io
import uuid
from datetime import date, datetime

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.bank_account import BankAccount
from app.models.bank_transaction import BankTransaction
from app.models.expense import Expense
from app.schemas.bank import (
    BankAccountCreate,
    BankAccountResponse,
    BankAccountUpdate,
    BankTransactionCreate,
    BankTransactionResponse,
    ImportStatementResponse,
    ReconcileRequest,
)


def _tx_response(tx: BankTransaction) -> BankTransactionResponse:
    data = {c.name: getattr(tx, c.name) for c in tx.__table__.columns}
    data["bank_account_alias"] = tx.bank_account.alias if tx.bank_account else None
    data["reconciled_expense_desc"] = tx.reconciled_expense.description if tx.reconciled_expense else None
    return BankTransactionResponse(**data)


def create_account(db: Session, data: BankAccountCreate) -> BankAccountResponse:
    account = BankAccount(**data.model_dump())
    db.add(account)
    db.commit()
    db.refresh(account)
    return BankAccountResponse.model_validate(account)


def get_accounts(db: Session) -> list[BankAccountResponse]:
    accounts = db.query(BankAccount).order_by(BankAccount.alias).all()
    return [BankAccountResponse.model_validate(a) for a in accounts]


def get_account(db: Session, account_id: uuid.UUID) -> BankAccountResponse:
    account = db.query(BankAccount).filter(BankAccount.id == account_id).first()
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cuenta bancaria no encontrada.")
    return BankAccountResponse.model_validate(account)


def update_account(db: Session, account_id: uuid.UUID, data: BankAccountUpdate) -> BankAccountResponse:
    account = db.query(BankAccount).filter(BankAccount.id == account_id).first()
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cuenta bancaria no encontrada.")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(account, field, value)
    db.commit()
    db.refresh(account)
    return BankAccountResponse.model_validate(account)


def create_transaction(db: Session, data: BankTransactionCreate) -> BankTransactionResponse:
    account = db.query(BankAccount).filter(BankAccount.id == data.bank_account_id).first()
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cuenta bancaria no encontrada.")

    tx = BankTransaction(**data.model_dump())
    db.add(tx)

    if data.transaction_type == "credit":
        account.balance = float(account.balance) + data.amount
    else:
        account.balance = float(account.balance) - data.amount

    db.commit()
    db.refresh(tx)
    return _tx_response(tx)


def get_transactions(
    db: Session,
    bank_account_id: uuid.UUID | None = None,
    reconciled: bool | None = None,
    page: int = 1,
    page_size: int = 50,
) -> tuple[list[BankTransactionResponse], int]:
    query = db.query(BankTransaction)
    if bank_account_id:
        query = query.filter(BankTransaction.bank_account_id == bank_account_id)
    if reconciled is not None:
        query = query.filter(BankTransaction.reconciled == reconciled)

    total = query.count()
    txs = query.order_by(BankTransaction.transaction_date.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return [_tx_response(t) for t in txs], total


def reconcile(db: Session, data: ReconcileRequest) -> BankTransactionResponse:
    tx = db.query(BankTransaction).filter(BankTransaction.id == data.transaction_id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Movimiento bancario no encontrado.")
    if tx.reconciled:
        raise HTTPException(status_code=400, detail="Este movimiento ya esta conciliado.")

    expense = db.query(Expense).filter(Expense.id == data.expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="Gasto no encontrado.")
    if expense.status != "approved":
        raise HTTPException(status_code=400, detail="Solo se pueden conciliar gastos aprobados.")

    tx.reconciled = True
    tx.reconciled_expense_id = expense.id
    db.commit()
    db.refresh(tx)
    return _tx_response(tx)


def unreconcile(db: Session, transaction_id: uuid.UUID) -> BankTransactionResponse:
    tx = db.query(BankTransaction).filter(BankTransaction.id == transaction_id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Movimiento bancario no encontrado.")
    if not tx.reconciled:
        raise HTTPException(status_code=400, detail="Este movimiento no esta conciliado.")

    tx.reconciled = False
    tx.reconciled_expense_id = None
    db.commit()
    db.refresh(tx)
    return _tx_response(tx)


def get_reconciliation_summary(db: Session, bank_account_id: uuid.UUID | None = None) -> dict:
    query = db.query(BankTransaction)
    if bank_account_id:
        query = query.filter(BankTransaction.bank_account_id == bank_account_id)

    total = query.count()
    reconciled_count = query.filter(BankTransaction.reconciled == True).count()  # noqa: E712
    pending_count = total - reconciled_count

    return {
        "total_transactions": total,
        "reconciled": reconciled_count,
        "pending": pending_count,
        "reconciliation_percentage": round((reconciled_count / total) * 100, 1) if total > 0 else 0,
    }


def _parse_money(value) -> float:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace("$", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return 0


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "date"):
        return value.date()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _find_column(headers: list[str], candidates: list[str]) -> int | None:
    for i, h in enumerate(headers):
        normalized = h.lower().strip()
        for c in candidates:
            if c in normalized:
                return i
    return None


def _parse_rows(rows: list[list]) -> tuple[list[dict], list[str]]:
    if not rows or len(rows) < 2:
        return [], ["Archivo vacío o sin datos."]

    headers = [str(h or "").strip() for h in rows[0]]

    date_col = _find_column(headers, ["fecha", "date"])
    desc_col = _find_column(headers, ["descripcion", "descripción", "description", "detalle", "glosa"])
    amount_col = _find_column(headers, ["monto", "amount", "valor"])
    cargo_col = _find_column(headers, ["cargo", "débito", "debito", "debit"])
    abono_col = _find_column(headers, ["abono", "crédito", "credito", "credit", "haber"])
    ref_col = _find_column(headers, ["referencia", "reference", "ref", "n° documento", "numero"])

    if date_col is None:
        return [], ["No se encontró columna de fecha."]
    if amount_col is None and cargo_col is None:
        return [], ["No se encontró columna de monto, cargo o abono."]

    parsed = []
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            tx_date = _parse_date(row[date_col] if date_col < len(row) else None)
            if not tx_date:
                errors.append(f"Fila {row_idx}: fecha inválida")
                continue

            description = str(row[desc_col]).strip() if desc_col is not None and desc_col < len(row) and row[desc_col] else ""
            reference = str(row[ref_col]).strip() if ref_col is not None and ref_col < len(row) and row[ref_col] else None

            if amount_col is not None:
                raw_amount = _parse_money(row[amount_col] if amount_col < len(row) else 0)
                if raw_amount >= 0:
                    tx_type = "credit"
                    amount = raw_amount
                else:
                    tx_type = "debit"
                    amount = abs(raw_amount)
            else:
                cargo = _parse_money(row[cargo_col] if cargo_col is not None and cargo_col < len(row) else 0)
                abono = _parse_money(row[abono_col] if abono_col is not None and abono_col < len(row) else 0)
                if cargo > 0:
                    tx_type = "debit"
                    amount = cargo
                elif abono > 0:
                    tx_type = "credit"
                    amount = abono
                else:
                    errors.append(f"Fila {row_idx}: sin monto")
                    continue

            if amount == 0:
                errors.append(f"Fila {row_idx}: monto cero")
                continue

            parsed.append({
                "transaction_date": tx_date,
                "description": description,
                "reference": reference,
                "amount": amount,
                "transaction_type": tx_type,
            })
        except Exception as e:
            errors.append(f"Fila {row_idx}: {e}")

    return parsed, errors


def _read_csv(content: bytes) -> list[list]:
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        return []

    for delimiter in (";", ",", "\t"):
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = [row for row in reader if any(cell.strip() for cell in row)]
        if rows and len(rows[0]) >= 2:
            return rows
    return []


def _read_excel(content: bytes) -> list[list]:
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def import_bank_statement(db: Session, file: UploadFile, bank_account_id: uuid.UUID) -> ImportStatementResponse:
    account = db.query(BankAccount).filter(BankAccount.id == bank_account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Cuenta bancaria no encontrada.")

    content = file.file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".csv") or filename.endswith(".txt"):
        rows = _read_csv(content)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        rows = _read_excel(content)
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado. Use CSV o Excel (.xlsx).")

    parsed, errors = _parse_rows(rows)

    total_credit = 0.0
    total_debit = 0.0
    imported = 0

    for item in parsed:
        tx = BankTransaction(
            bank_account_id=bank_account_id,
            transaction_date=item["transaction_date"],
            amount=item["amount"],
            transaction_type=item["transaction_type"],
            reference=item["reference"],
            description=item["description"],
        )
        db.add(tx)

        if item["transaction_type"] == "credit":
            total_credit += item["amount"]
            account.balance = float(account.balance) + item["amount"]
        else:
            total_debit += item["amount"]
            account.balance = float(account.balance) - item["amount"]

        imported += 1

    db.commit()

    return ImportStatementResponse(
        imported=imported,
        skipped=len(errors),
        errors=errors[:20],
        total_credit=total_credit,
        total_debit=total_debit,
    )
