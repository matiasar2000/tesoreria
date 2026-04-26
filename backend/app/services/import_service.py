import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.budget_item import BudgetItem
from app.models.expense import Expense
from app.models.fiscal_year import FiscalYear


def _parse_money(value) -> float:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace("$", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return 0


def import_budget_from_excel(db: Session, file: UploadFile, fiscal_year_id: uuid.UUID) -> dict:
    fy = db.query(FiscalYear).filter(FiscalYear.id == fiscal_year_id).first()
    if not fy:
        raise HTTPException(status_code=404, detail="Año fiscal no encontrado.")

    content = file.file.read()
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    rows_processed = 0
    rows_error = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            number = int(row[0])
            name = str(row[1]).strip()
            authority = str(row[2]).strip().lower() if row[2] else "mixto"
            allocated = _parse_money(row[3])

            existing = db.query(BudgetItem).filter(
                BudgetItem.fiscal_year_id == fiscal_year_id,
                BudgetItem.number == number,
            ).first()

            if existing:
                existing.name = name
                existing.authority = authority
                existing.allocated_amount = allocated
            else:
                db.add(BudgetItem(
                    fiscal_year_id=fiscal_year_id,
                    number=number,
                    name=name,
                    authority=authority,
                    allocated_amount=allocated,
                ))
            rows_processed += 1
        except Exception as e:
            rows_error += 1
            errors.append(f"Fila {row_idx}: {e}")

    db.commit()
    return {"rows_processed": rows_processed, "rows_error": rows_error, "errors": errors}


def import_expenses_from_excel(db: Session, file: UploadFile, fiscal_year_id: uuid.UUID, user_id: uuid.UUID) -> dict:
    fy = db.query(FiscalYear).filter(FiscalYear.id == fiscal_year_id).first()
    if not fy:
        raise HTTPException(status_code=404, detail="Año fiscal no encontrado.")

    content = file.file.read()
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    rows_processed = 0
    rows_error = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            item_number = int(row[0])
            amount = _parse_money(row[1])
            description = str(row[2]).strip() if row[2] else "Importado desde Excel"
            expense_date_val = row[3] if row[3] else date.today()
            if isinstance(expense_date_val, str):
                expense_date_val = date.fromisoformat(expense_date_val)
            elif hasattr(expense_date_val, "date"):
                expense_date_val = expense_date_val.date()
            supplier = str(row[4]).strip() if len(row) > 4 and row[4] else None

            item = db.query(BudgetItem).filter(
                BudgetItem.fiscal_year_id == fiscal_year_id,
                BudgetItem.number == item_number,
            ).first()
            if not item:
                raise ValueError(f"Partida {item_number} no encontrada.")

            expense = Expense(
                budget_item_id=item.id,
                fiscal_year_id=fiscal_year_id,
                amount=amount,
                description=description,
                expense_date=expense_date_val,
                supplier_name=supplier,
                status="approved",
                requested_by_id=user_id,
                approved_by_id=user_id,
            )
            db.add(expense)

            item.executed_amount = Decimal(str(item.executed_amount)) + Decimal(str(amount))
            rows_processed += 1
        except Exception as e:
            rows_error += 1
            errors.append(f"Fila {row_idx}: {e}")

    db.commit()
    return {"rows_processed": rows_processed, "rows_error": rows_error, "errors": errors}
