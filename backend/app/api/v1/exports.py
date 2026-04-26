import uuid
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.database import get_db
from app.models.user import User
from app.schemas.budget_item import BudgetItemResponse
from app.services.budget_service import get_budget_items
from app.services.expense_service import get_expenses

router = APIRouter(prefix="/exports", tags=["Exportaciones"])

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
AMOUNT_FORMAT = '#,##0'
PERCENT_FORMAT = "0.0%"
DATE_FORMAT = "dd-mm-yyyy"

BUDGET_STATUS_LABELS = {
    "green": "Verde",
    "yellow": "Amarillo",
    "red": "Rojo",
}

EXPENSE_STATUS_LABELS = {
    "draft": "Borrador",
    "pending_review": "En revisión",
    "pending_approval": "Pendiente",
    "pending_directorio": "Pendiente directorio",
    "approved": "Aprobado",
    "rejected": "Rechazado",
    "voided": "Anulado",
    "paid": "Pagado",
}


def _format_header(worksheet: Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _auto_adjust_widths(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def _build_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _to_number(value: object) -> float:
    return float(value or 0)


@router.get("/budget")
def export_budget(
    fiscal_year_id: uuid.UUID,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Presupuesto"

    headers = [
        "N°",
        "Nombre",
        "Autoridad",
        "Monto Asignado",
        "Monto Ejecutado",
        "Disponible",
        "% Ejecución",
        "Estado",
    ]
    worksheet.append(headers)

    for item in get_budget_items(db, fiscal_year_id):
        response = BudgetItemResponse.model_validate(item)
        worksheet.append(
            [
                response.number,
                response.name,
                response.authority,
                _to_number(response.allocated_amount),
                _to_number(response.executed_amount),
                _to_number(response.available_amount),
                _to_number(response.execution_percentage) / 100,
                BUDGET_STATUS_LABELS.get(response.status_color, response.status_color),
            ]
        )

    for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = AMOUNT_FORMAT

    for cell in worksheet.iter_cols(min_col=7, max_col=7, min_row=2):
        for percentage_cell in cell:
            percentage_cell.number_format = PERCENT_FORMAT

    _format_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _auto_adjust_widths(worksheet)

    return _build_response(workbook, f"presupuesto_{fiscal_year_id}.xlsx")


@router.get("/expenses")
def export_expenses(
    fiscal_year_id: uuid.UUID,
    status: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    _, total = get_expenses(db, fiscal_year_id, None, status, page=1, page_size=1)
    expenses, _ = get_expenses(db, fiscal_year_id, None, status, page=1, page_size=max(total, 1))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Gastos"

    headers = [
        "Fecha",
        "Descripción",
        "Partida",
        "Proveedor",
        "RUT",
        "N° Factura",
        "Monto",
        "Estado",
        "Solicitado por",
    ]
    worksheet.append(headers)

    for expense in expenses:
        worksheet.append(
            [
                expense.expense_date,
                expense.description,
                expense.budget_item_name or "",
                expense.supplier_name or "",
                expense.supplier_rut or "",
                expense.invoice_number or "",
                _to_number(expense.amount),
                EXPENSE_STATUS_LABELS.get(expense.status, expense.status),
                expense.requested_by_name or "",
            ]
        )

    for date_cell in worksheet.iter_cols(min_col=1, max_col=1, min_row=2):
        for cell in date_cell:
            cell.number_format = DATE_FORMAT

    for amount_cell in worksheet.iter_cols(min_col=7, max_col=7, min_row=2):
        for cell in amount_cell:
            cell.number_format = AMOUNT_FORMAT

    _format_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _auto_adjust_widths(worksheet)

    return _build_response(workbook, f"gastos_{fiscal_year_id}.xlsx")
