from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, get_db
from app.models.user import User
from app.schemas.reports import AnnualBalance, QuarterlyBalance, ReportsSummary
from app.services import report_service

router = APIRouter(prefix="/reports", tags=["reports"])

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
AMOUNT_FORMAT = '#,##0'
PERCENT_FORMAT = "0.0%"


def _format_header(worksheet: Worksheet, row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for cell in worksheet[row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _auto_adjust_widths(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def _build_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/summary", response_model=ReportsSummary)
def get_reports_summary(
    year: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return report_service.get_reports(db, year=year)


@router.get("/quarterly", response_model=QuarterlyBalance)
def get_quarterly_balance(
    year: int | None = Query(None),
    quarter: int = Query(..., ge=1, le=4),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return report_service.get_quarterly_balance(db, year=year, quarter=quarter)


@router.get("/annual", response_model=AnnualBalance)
def get_annual_balance(
    year: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return report_service.get_annual_balance(db, year=year)


@router.get("/annual/export")
def export_annual_balance(
    year: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    annual = report_service.get_annual_balance(db, year=year)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary.append(["Indicador", "Valor"])
    summary.append(["Ano fiscal", annual.fiscal_year])
    summary.append(["Presupuesto total", annual.total_budget])
    summary.append(["Ingresos", annual.total_income])
    summary.append(["Gastos aprobados", annual.total_expenses])
    summary.append(["Saldo final", annual.final_balance])
    summary.append(["Ejecucion", annual.execution_percentage / 100])
    summary.append(["Gastos pendientes", annual.pending_expenses])
    summary.append(["Gastos aprobados", annual.approved_expenses])
    summary.append(["Gastos anulados", annual.voided_expenses])
    for row in summary.iter_rows(min_row=3, max_row=6, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = AMOUNT_FORMAT
    summary["B7"].number_format = PERCENT_FORMAT
    _format_header(summary)
    _auto_adjust_widths(summary)

    quarters = workbook.create_sheet("Trimestres")
    quarters.append([
        "Trimestre",
        "Periodo",
        "Presupuesto",
        "Ingresos",
        "Gastos",
        "Saldo",
        "% Ejecucion",
    ])
    for quarter in annual.quarterly_summary:
        quarters.append([
            quarter.quarter_label,
            f"{quarter.period_start} / {quarter.period_end}",
            quarter.total_budget,
            quarter.total_income,
            quarter.total_expenses,
            quarter.balance,
            quarter.execution_percentage / 100,
        ])
    for row in quarters.iter_rows(min_row=2, min_col=3, max_col=6):
        for cell in row:
            cell.number_format = AMOUNT_FORMAT
    for row in quarters.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            cell.number_format = PERCENT_FORMAT
    _format_header(quarters)
    quarters.freeze_panes = "A2"
    _auto_adjust_widths(quarters)

    items = workbook.create_sheet("Partidas")
    items.append([
        "N",
        "Partida",
        "Asignado",
        "Ejecutado",
        "Disponible",
        "% Ejecucion",
        "Estado",
    ])
    for item in annual.expenses_by_item:
        items.append([
            item.item_number,
            item.item_name,
            item.allocated,
            item.executed,
            item.available,
            item.percentage / 100,
            item.status_color,
        ])
    for row in items.iter_rows(min_row=2, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = AMOUNT_FORMAT
    for row in items.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = PERCENT_FORMAT
    _format_header(items)
    items.freeze_panes = "A2"
    items.auto_filter.ref = items.dimensions
    _auto_adjust_widths(items)

    banks = workbook.create_sheet("Cuentas Bancarias")
    banks.append(["Cuenta", "Saldo"])
    for account in annual.bank_balances:
        banks.append([account.account_name, account.balance])
    for row in banks.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = AMOUNT_FORMAT
    _format_header(banks)
    banks.freeze_panes = "A2"
    _auto_adjust_widths(banks)

    return _build_response(workbook, f"balance_anual_{annual.fiscal_year}.xlsx")
